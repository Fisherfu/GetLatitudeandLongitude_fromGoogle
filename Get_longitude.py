# -*- coding: utf-8 -*-
"""
Created on Thu Aug  3 09:13:39 2023

@author: USER
"""

#放在第一個cell
import codecs
import os
os.chdir("C:/Users/USER/Downloads")
f = codecs.open("test_data.txt", 'r', 'utf-8') #記得以utf-8開啟文字檔
location = f.read()
f.close()

#放在第二個cell
places = tuple(location.split('\n'))
URL = []
for i in places:
    URL.append("https://www.google.com/maps/place?q=" + i)
    
    
#放在第三個cell
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'longitude'
ws['B1'] = 'latitude'    

#放在第四個cell
def STR_to_NUM(data):
    line = tuple(data.split(',')) #註1
    num1 = float(line[1])
    num2 = float(line[2])
    line = [num1, num2]
    return line

#先放在第五個cell，之後刪除
import requests
from bs4 import BeautifulSoup
response = requests.get(URL[0])
soup = BeautifulSoup(response.text, "html.parser")
text = soup.prettify()
print(text)

#放在第五個cell，記得先將之前寫的刪除
def coordination(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    text = soup.prettify() #text 包含了html的內容
    initial_pos = text.find(";window.APP_INITIALIZATION_STATE")
    #尋找;window.APP_INITIALIZATION_STATE所在位置
    data = text[initial_pos+36:initial_pos+85] #將其後的參數進行存取
    num_data = STR_to_NUM(data)
    ws.append(num_data) #將經緯度存到Excel裡


def STR_to_NUM(data):
    line = tuple(data.split(','))
    num1 = float(line[1])
    num2 = float(line[2])
    line = [num1, num2]
    return line

#放在第六個cell
import requests
from bs4 import BeautifulSoup
for i in URL:
    coordination(i)
    wb.save('test.xlsx') #自行決定檔名
